import numpy as np
import openpyxl

table=np.zeros((40,4,5),int)

wb_obj = openpyxl.load_workbook("TPP.xlsx")
sheet_obj = wb_obj.active
row_count = sheet_obj.max_row

letter_dict = {
  "M": 0,
  "L": 1,
  "H": 2,
  "Q": 3
}

for i in range(2, row_count+1):
    s = sheet_obj.cell(row=i, column=1).value
    ch = '-'
    index = s.find(ch)
    index1=int(s[0:index]) - 1
    letter = s[len(s) - 1]
    index2=int(letter_dict[letter])

    table[index1][index2][0] = sheet_obj.cell(row=i, column=3).value if (sheet_obj.cell(row=i, column=3).value is not None) else 0
    table[index1][index2][1] = sheet_obj.cell(row=i, column=4).value if (sheet_obj.cell(row=i, column=4).value is not None) else 0
    table[index1][index2][2] = sheet_obj.cell(row=i, column=5).value if (sheet_obj.cell(row=i, column=5).value is not None) else 0
    table[index1][index2][3] = sheet_obj.cell(row=i, column=6).value if (sheet_obj.cell(row=i, column=6).value is not None) else 0
    table[index1][index2][4] = sheet_obj.cell(row=i, column=7).value if (sheet_obj.cell(row=i, column=7).value is not None) else 0

    # print(table[index1][index2][0])
    # print(table[index1][index2][1])
    # print(table[index1][index2][2])
    # print(table[index1][index2][3])
    # print(table[index1][index2][4])
