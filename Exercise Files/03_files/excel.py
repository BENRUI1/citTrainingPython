import openpyxl
from openpyxl import Workbook
import os

def main():
    #ask file name - enter file name without extension
    filename = input("Enter filename : ")

    # try..except to check if file exists and create it not existing
    # if file exists
    #       ask if file should be deleted. If answer is Y, delete file and set operation to write
    #       else set operation to read

    try:
        wb_obj = openpyxl.load_workbook(filename+".xlsx")
        v_del_file = input("File exists, do you want to replace it? (Y/N)")
        if v_del_file == "Y":
            os.remove(filename+".xlsx")
            v_operation = "write"
        else:
            v_operation="read"
    except:
        v_operation="write"

    if v_operation == "write":
        wb_obj = Workbook()
        wb_obj.save(filename+".xlsx")

    sheet_obj = wb_obj.active
    #If file exists, read it
    if v_operation == "read":
        row_count = sheet_obj.max_row
        for line in range(1, row_count):
            cell_obj = sheet_obj.cell(row = line, column = 1)
            print(cell_obj.value)
    #if file does not exist, create file
    else:
        line = 0
        name = True
        while name:
            line = line + 1
            name = input("Enter name : ")
            cell_obj = sheet_obj.cell(row = line, column = 1)
            cell_obj.value = name

        wb_obj.save(filename + ".xlsx")


if __name__ == "__main__":
    main()